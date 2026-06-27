#!/usr/bin/env python3
# =========================================
# ID: PHYLOFACE_CAPITALS_DISTANCE_MATRIX
# VERSION: v0.1
# =========================================
# FILE: scripts/build_capitals_distance_matrix.py
#
# Arma una matriz de distancia great-circle (haversine, km) entre las capitales
# de todos los países soberanos, enriquecida con:
#   - idioma(s) oficiales        (mledoze/countries.json)
#   - último colonizador europeo (OWID / COLDAT de Becker)
#   - flag de selección del Mundial 2026
# Pensado como features geográfico-culturales para analizar patrones migratorios
# detrás del parecido facial entre selecciones (vitrina).
#
# Fuentes (se cachean en data/input/geo/ para reproducibilidad):
#   - capitales+coords+población: gist ofou country-capital-lat-long-population.csv
#   - idiomas/región/ISO3/alias:  mledoze/countries.json
#   - colonizador:                OWID european-overseas-colonies-and-their-last-colonizer.csv
#
# Salidas (data/output/geo/):
#   - world_capitals_distance.json   (canónico: countries[] + matrix km)
#   - world_capitals.xlsx            (hoja Países + hoja Matriz)
#   - world_capitals_pairs.csv       (long-format, ordenado por distancia)
#
# Uso:
#   conda run -n face-sim python scripts/build_capitals_distance_matrix.py
#   ... --refresh   # ignora cache y re-descarga fuentes

import argparse
import csv
import json
import math
import os
import re
import sys
import unicodedata

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
RAW_DIR = os.path.join(REPO_ROOT, "data", "input", "geo")
OUT_DIR = os.path.join(REPO_ROOT, "data", "output", "geo")
UA = "Mozilla/5.0 (X11; Linux x86_64) Chrome/124.0 Safari/537.36"

SRC = {
    "capitals": ("country-capital-lat-long-population.csv",
                 "https://gist.githubusercontent.com/ofou/df09a6834a8421b4f376c875194915c9/raw/country-capital-lat-long-population.csv"),
    "countries": ("mledoze_countries.json",
                  "https://raw.githubusercontent.com/mledoze/countries/master/countries.json"),
    "colonizer": ("owid_last_colonizer.csv",
                  "https://ourworldindata.org/grapher/european-overseas-colonies-and-their-last-colonizer.csv?csvType=full"),
}

# 48 selecciones del Mundial 2026 (nombre común en inglés -> se resuelve a cca3).
# England/Scotland no son estados soberanos en los datasets de capitales: caen en
# United Kingdom (GBR) y quedan marcadas vía esa fila.
WC2026 = [
    "Germany", "Saudi Arabia", "Algeria", "Argentina", "Australia", "Austria",
    "Bosnia and Herzegovina", "Brazil", "Belgium", "Canada", "Qatar",
    "Czech Republic", "Colombia", "Ivory Coast", "Croatia", "Curaçao",
    "United States", "Ecuador", "Egypt", "United Kingdom",  # England/Scotland -> UK
    "Spain", "France", "Ghana", "Haiti", "Iran", "Iraq", "Cape Verde", "Japan",
    "Jordan", "Morocco", "Mexico", "Norway", "New Zealand", "Panama", "Paraguay",
    "Netherlands", "Portugal", "DR Congo", "South Korea", "Senegal",
    "South Africa", "Sweden", "Switzerland", "Turkey", "Tunisia", "Uruguay",
    "Uzbekistan",
]

# Overrides nombre(capitales CSV) -> cca3, para los que el match automático falla.
NAME_OVERRIDES = {
    "united states": "USA", "south korea": "KOR", "north korea": "PRK",
    "russia": "RUS", "iran": "IRN", "syria": "SYR", "laos": "LAO",
    "vietnam": "VNM", "brunei": "BRN", "moldova": "MDA", "tanzania": "TZA",
    "venezuela": "VEN", "bolivia": "BOL", "ivory coast": "CIV",
    "cote d'ivoire": "CIV", "cape verde": "CPV", "dr congo": "COD",
    "democratic republic of the congo": "COD", "republic of the congo": "COG",
    "congo": "COG", "the gambia": "GMB", "gambia": "GMB", "czechia": "CZE",
    "czech republic": "CZE", "swaziland": "SWZ", "eswatini": "SWZ",
    "macedonia": "MKD", "north macedonia": "MKD", "burma": "MMR",
    "myanmar": "MMR", "east timor": "TLS", "timor-leste": "TLS",
    "vatican city": "VAT", "vatican": "VAT", "palestine": "PSE",
    "kosovo": "XKX", "turkey": "TUR", "turkiye": "TUR",
    "dem people s republic of korea": "PRK",
    "tfyr macedonia": "MKD", "taiwan province of china": "TWN", "taiwan": "TWN",
    "china hong kong sar": "HKG", "china macao sar": "MAC",
    "china taiwan province of china": "TWN",
    "brunei darussalam": "BRN", "the former yugoslav republic of macedonia": "MKD",
}


def norm(s):
    s = unicodedata.normalize("NFKD", s or "").encode("ascii", "ignore").decode()
    s = re.sub(r"\(.*?\)", "", s)           # quita paréntesis "Tiranë (Tirana)"
    s = re.sub(r"[^a-z0-9 ]", " ", s.lower())
    return re.sub(r"\s+", " ", s).strip()


def fetch(name, refresh):
    fname, url = SRC[name]
    os.makedirs(RAW_DIR, exist_ok=True)
    path = os.path.join(RAW_DIR, fname)
    if os.path.exists(path) and not refresh:
        return open(path, "rb").read()
    r = requests.get(url, headers={"User-Agent": UA, "Accept": "text/csv,application/json"},
                     timeout=60)
    r.raise_for_status()
    with open(path, "wb") as f:
        f.write(r.content)
    return r.content


def load_countries(refresh):
    """mledoze -> name_norm -> {cca3, common, languages, region, subregion}."""
    data = json.loads(fetch("countries", refresh))
    by_cca3, name_to_cca3 = {}, {}
    for c in data:
        cca3 = c.get("cca3")
        rec = {
            "cca3": cca3,
            "common": c["name"]["common"],
            "official": c["name"].get("official"),
            "languages": sorted((c.get("languages") or {}).values()),
            "region": c.get("region"),
            "subregion": c.get("subregion"),
        }
        by_cca3[cca3] = rec
        names = {c["name"]["common"], c["name"].get("official", "")}
        names |= set(c.get("altSpellings") or [])
        for tr in (c.get("translations") or {}).values():
            names.add(tr.get("common", ""))
        for n in names:
            k = norm(n)
            if k:
                name_to_cca3.setdefault(k, cca3)
    return by_cca3, name_to_cca3


def load_colonizer(refresh):
    """OWID -> cca3 -> last colonizer (limpia el 'zzzz. Never colonized')."""
    txt = fetch("colonizer", refresh).decode("utf-8")
    out = {}
    for row in csv.DictReader(txt.splitlines()):
        code = row.get("Code")
        col = (row.get("Last colonizer") or "").strip()
        low = col.lower()
        if "never colonized" in low or low.startswith("zzzz"):
            col = "Never colonized"
        elif low.startswith("zz") and "coloniz" in low:   # "zz. Colonizer" = potencia colonizadora
            col = "Colonizer power"
        elif low.startswith("z") and "multiple" in low:   # "z. Multiple colonizers"
            col = "Multiple colonizers"
        if code:
            out[code] = col
    return out


def load_capitals(refresh):
    txt = fetch("capitals", refresh).decode("utf-8")
    rows = []
    for r in csv.DictReader(txt.splitlines()):
        try:
            lat = float(r["Latitude"]); lon = float(r["Longitude"])
        except (ValueError, KeyError):
            continue
        rows.append({
            "country_raw": r["Country"].strip(),
            "capital": re.sub(r"\s*\(.*?\)\s*", "", r["Capital City"]).strip(),
            "lat": lat, "lon": lon,
            "population": int(r["Population"]) if (r.get("Population") or "").strip().isdigit() else None,
        })
    return rows


def haversine(a, b):
    R = 6371.0
    lat1, lon1, lat2, lon2 = map(math.radians, (a["lat"], a["lon"], b["lat"], b["lon"]))
    dlat, dlon = lat2 - lat1, lon2 - lon1
    h = math.sin(dlat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon / 2) ** 2
    return 2 * R * math.asin(math.sqrt(h))


def main():
    ap = argparse.ArgumentParser(description="Matriz de distancia entre capitales + colonizador/idioma.")
    ap.add_argument("--refresh", action="store_true", help="re-descarga fuentes (ignora cache).")
    args = ap.parse_args()

    print("[geo] cargando fuentes...")
    by_cca3, name_to_cca3 = load_countries(args.refresh)
    colonizer = load_colonizer(args.refresh)
    caps = load_capitals(args.refresh)

    wc_cca3 = {name_to_cca3.get(norm(n)) or NAME_OVERRIDES.get(norm(n)) for n in WC2026}
    wc_cca3.discard(None)

    countries, unmatched = [], []
    for c in caps:
        k = norm(c["country_raw"])
        cca3 = NAME_OVERRIDES.get(k) or name_to_cca3.get(k)
        meta = by_cca3.get(cca3, {})
        if not cca3:
            unmatched.append(c["country_raw"])
        countries.append({
            "country": meta.get("common") or c["country_raw"],
            "country_raw": c["country_raw"],
            "cca3": cca3,
            "capital": c["capital"],
            "lat": c["lat"], "lon": c["lon"],
            "population": c["population"],
            "region": meta.get("region"),
            "subregion": meta.get("subregion"),
            "languages": meta.get("languages") or [],
            "last_colonizer": colonizer.get(cca3),
            "is_wc2026": cca3 in wc_cca3 if cca3 else False,
        })

    # Deduplicar por cca3 (algún dataset trae territorios repetidos); ordenar.
    seen, dedup = set(), []
    for c in sorted(countries, key=lambda x: x["country"]):
        key = c["cca3"] or c["country_raw"]
        if key in seen:
            continue
        seen.add(key)
        dedup.append(c)
    countries = dedup
    n = len(countries)
    print(f"[geo] {n} capitales | sin match a ISO3: {len(unmatched)} "
          f"({', '.join(unmatched[:8])}{'...' if len(unmatched) > 8 else ''})")
    print(f"[geo] colonizador resuelto: {sum(1 for c in countries if c['last_colonizer'])}/{n} | "
          f"idioma: {sum(1 for c in countries if c['languages'])}/{n} | "
          f"mundial2026: {sum(1 for c in countries if c['is_wc2026'])}")

    # Matriz haversine.
    matrix = [[0.0] * n for _ in range(n)]
    for i in range(n):
        for j in range(i + 1, n):
            d = round(haversine(countries[i], countries[j]), 1)
            matrix[i][j] = matrix[j][i] = d

    os.makedirs(OUT_DIR, exist_ok=True)

    # --- JSON canónico ---
    jpath = os.path.join(OUT_DIR, "world_capitals_distance.json")
    with open(jpath, "w", encoding="utf-8") as f:
        json.dump({
            "schema": "phyloface-capitals-distance-v0.1",
            "distance": "great-circle haversine, km",
            "countries_count": n,
            "unmatched_iso3": unmatched,
            "countries": countries,
            "matrix_km": matrix,
        }, f, ensure_ascii=False)
    print(f"[geo] JSON -> {jpath}")

    # --- Excel (Países + Matriz) ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Países"
    cols = [("País", "country", 22), ("ISO3", "cca3", 8), ("Capital", "capital", 20),
            ("Lat", "lat", 10), ("Lon", "lon", 10), ("Población", "population", 12),
            ("Región", "region", 14), ("Subregión", "subregion", 18),
            ("Idiomas", "languages", 30), ("Último colonizador", "last_colonizer", 18),
            ("Mundial 2026", "is_wc2026", 12)]
    hf, hfont = PatternFill("solid", fgColor="1F3864"), Font(bold=True, color="FFFFFF")
    for c, (label, _k, w) in enumerate(cols, 1):
        cell = ws.cell(1, c, label); cell.fill = hf; cell.font = hfont
        ws.column_dimensions[get_column_letter(c)].width = w
    for r, country in enumerate(countries, 2):
        for c, (_l, key, _w) in enumerate(cols, 1):
            v = country[key]
            if isinstance(v, list):
                v = ", ".join(v)
            if key == "is_wc2026":
                v = "Sí" if v else ""
            ws.cell(r, c, v)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{n+1}"

    wm = wb.create_sheet("Matriz")
    wm.cell(1, 1, "Capital \\ Capital").font = hfont
    wm.cell(1, 1).fill = hf
    for j, country in enumerate(countries, 2):
        hc = wm.cell(1, j, country["capital"]); hc.fill = hf; hc.font = hfont
        rc = wm.cell(j, 1, country["capital"]); rc.fill = hf; rc.font = hfont
    for i in range(n):
        for j in range(n):
            wm.cell(i + 2, j + 2, matrix[i][j])
    wm.freeze_panes = "B2"

    xpath = os.path.join(OUT_DIR, "world_capitals.xlsx")
    wb.save(xpath)
    print(f"[geo] XLSX -> {xpath}")

    # --- Pairs CSV (long, ordenado por distancia) ---
    pairs = []
    for i in range(n):
        for j in range(i + 1, n):
            pairs.append((countries[i], countries[j], matrix[i][j]))
    pairs.sort(key=lambda p: p[2])
    cpath = os.path.join(OUT_DIR, "world_capitals_pairs.csv")
    with open(cpath, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["pais_a", "capital_a", "pais_b", "capital_b", "km",
                    "mismo_colonizador", "comparten_idioma", "ambos_wc2026"])
        for a, b, d in pairs:
            same_col = bool(a["last_colonizer"] and a["last_colonizer"] == b["last_colonizer"]
                            and a["last_colonizer"] != "Never colonized")
            share_lang = bool(set(a["languages"]) & set(b["languages"]))
            w.writerow([a["country"], a["capital"], b["country"], b["capital"], d,
                        "sí" if same_col else "", "sí" if share_lang else "",
                        "sí" if (a["is_wc2026"] and b["is_wc2026"]) else ""])
    print(f"[geo] PAIRS CSV -> {cpath}  ({len(pairs)} pares)")
    print(f"[geo] más cercanas: {pairs[0][0]['capital']} ↔ {pairs[0][1]['capital']} = {pairs[0][2]} km")
    print(f"[geo] más lejanas:  {pairs[-1][0]['capital']} ↔ {pairs[-1][1]['capital']} = {pairs[-1][2]} km")


if __name__ == "__main__":
    sys.exit(main())
