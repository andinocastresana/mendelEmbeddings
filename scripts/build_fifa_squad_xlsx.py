#!/usr/bin/env python3
# =========================================
# ID: PHYLOFACE_FIFA_SQUAD_XLSX
# VERSION: v0.1
# =========================================
# FILE: scripts/build_fifa_squad_xlsx.py
#
# Construye un Excel (.xlsx) con la ficha de cada jugador del Mundial 2026 a
# partir del manifiesto JSON de build_fifa_squad_manifest.py. Una fila por
# jugador, con enlace (clickable) a la foto de mejor resolucion. Pensado para
# que el usuario lo use luego como insumo del scraping de imagenes.
#
# Hoja "Jugadores": todas las fichas. Hoja "Resumen": cobertura por seleccion.
#
# IMPORTANTE: las URLs apuntan a fotos copyright FIFA/Getty. Uso local; no
# publicar/redistribuir (ver license_status del manifiesto).
#
# Uso:
#   conda run -n face-sim python scripts/build_fifa_squad_xlsx.py
#   ... --input <manifest.json> --output <salida.xlsx>

import argparse
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
DEFAULT_IN = os.path.join(REPO_ROOT, "data", "output", "teams",
                          "manifest_fifa_northamerica2026_official.json")
DEFAULT_OUT = os.path.join(REPO_ROOT, "data", "output", "teams",
                           "fichas_fifa_northamerica2026.xlsx")

# (encabezado, key del jugador, ancho)
COLUMNS = [
    ("Selección", "_team", 20),
    ("IdTeam", "_id_team", 10),
    ("Jugador", "name", 26),
    ("Dorsal", "jersey_number", 8),
    ("Posición", "position", 14),
    ("Nacimiento", "birth_date", 12),
    ("Altura (cm)", "height_cm", 10),
    ("Peso (kg)", "weight_kg", 9),
    ("País", "country", 7),
    ("FIFA ID", "id_player", 10),
    ("Foto (mejor resolución)", "photo_url_best", 60),
    ("Foto (URL base)", "photo_base_url", 60),
]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")
LINK_FONT = Font(color="0563C1", underline="single")


def build_players_sheet(ws, manifest):
    ws.title = "Jugadores"
    # Encabezado.
    for c, (label, _key, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(c)].width = width

    r = 2
    for team in manifest.get("teams", []):
        tname = team.get("team_name")
        tid = team.get("id_team")
        for p in team.get("players", []):
            row = dict(p)
            row["_team"] = tname
            row["_id_team"] = tid
            for c, (_label, key, _w) in enumerate(COLUMNS, 1):
                val = row.get(key)
                cell = ws.cell(row=r, column=c, value=val)
                # Las dos columnas de URL: hyperlink clickable, texto = URL (scrapeable).
                if key in ("photo_url_best", "photo_base_url") and val:
                    cell.hyperlink = val
                    cell.font = LINK_FONT
            r += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{r-1}"
    return r - 2  # n filas de datos


def build_summary_sheet(ws, manifest):
    ws.title = "Resumen"
    headers = ["Selección", "IdTeam", "Jugadores", "Con foto"]
    widths = [22, 10, 12, 10]
    for c, (label, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        ws.column_dimensions[get_column_letter(c)].width = width
    r = 2
    for team in manifest.get("teams", []):
        ws.cell(row=r, column=1, value=team.get("team_name"))
        ws.cell(row=r, column=2, value=team.get("id_team"))
        ws.cell(row=r, column=3, value=team.get("players_count", len(team.get("players", []))))
        ws.cell(row=r, column=4, value=team.get("photos_count"))
        r += 1
    # Totales.
    tcell = ws.cell(row=r, column=1, value="TOTAL")
    tcell.font = Font(bold=True)
    ws.cell(row=r, column=3, value=manifest.get("players_count")).font = Font(bold=True)
    ws.cell(row=r, column=4, value=manifest.get("photos_count")).font = Font(bold=True)
    ws.freeze_panes = "A2"


def main():
    ap = argparse.ArgumentParser(description="Excel de fichas FIFA WC2026 desde el manifiesto.")
    ap.add_argument("--input", default=DEFAULT_IN)
    ap.add_argument("--output", default=DEFAULT_OUT)
    args = ap.parse_args()

    with open(args.input, encoding="utf-8") as f:
        manifest = json.load(f)

    wb = Workbook()
    n_rows = build_players_sheet(wb.active, manifest)
    build_summary_sheet(wb.create_sheet(), manifest)

    # Nota de licencia como propiedad del documento.
    wb.properties.description = (
        f"{manifest.get('license_note', '')} "
        f"license_status={manifest.get('license_status')}")

    os.makedirs(os.path.dirname(args.output), exist_ok=True)
    wb.save(args.output)
    print(f"[xlsx] OK -> {args.output}")
    print(f"[xlsx] filas (jugadores)={n_rows}  equipos={manifest.get('teams_count')}  "
          f"con_foto={manifest.get('photos_count')}")
    print(f"[xlsx] licencia: {manifest.get('license_status')} (publication_ok="
          f"{manifest.get('publication_ok')})")


if __name__ == "__main__":
    sys.exit(main())
